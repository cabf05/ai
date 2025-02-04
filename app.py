import os
import io
import logging
import traceback
import requests
import pytesseract
import nltk
from flask import Flask, request, redirect, session, render_template_string
from docx import Document
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from PIL import Image
import fitz  # PyMuPDF

# Configuração de ambiente
os.environ['NLTK_DATA'] = '/opt/render/nltk_data'
pytesseract.pytesseract.tesseract_cmd = '/usr/bin/tesseract'

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24))

# Serviços de IA
AI_SERVICES = {
    'OpenAI': {
        'guide': [
            '1. Acesse https://platform.openai.com/',
            '2. Crie uma API Key em "API Keys"',
            '3. Cole a chave abaixo'
        ],
        'api_url': 'https://api.openai.com/v1/chat/completions'
    },
    'HuggingFace': {
        'guide': [
            '1. Acesse https://huggingface.co/',
            '2. Crie um token de acesso com permissão "Read"',
            '3. Cole o token abaixo'
        ],
        'api_url': 'https://api-inference.huggingface.co/models/'
    },
    'Cohere': {
        'guide': [
            '1. Acesse https://dashboard.cohere.ai/',
            '2. Crie uma nova chave API',
            '3. Cole a chave gerada abaixo'
        ],
        'api_url': 'https://api.cohere.ai/v1/generate'
    }
}

HTML_BASE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Sistema de Resumo</title>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        .container {{ max-width: 800px; margin: 50px auto; }}
        .card {{ margin-top: 20px; padding: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        textarea {{ height: 400px; width: 100%; margin: 20px 0; }}
        .alert {{ margin: 20px 0; }}
        pre {{ white-space: pre-wrap; background: #f8f9fa; padding: 15px; }}
    </style>
</head>
<body>
    <div class="container">
        <h2 class="text-center mb-4">📄 Sistema de Resumo de Documentos</h2>
        <div class="card">
            <a href="/" class="btn btn-secondary mb-3">🏠 Voltar</a>
            {}
        </div>
    </div>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
'''

def extract_text_with_ocr(pdf_bytes):
    """Extrai texto de PDFs digitalizados usando OCR"""
    try:
        text = []
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        
        for page_num, page in enumerate(doc):
            try:
                pix = page.get_pixmap(dpi=300)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                text.append(pytesseract.image_to_string(img, lang='por+eng'))
            except Exception as e:
                logger.error(f"Erro na página {page_num + 1}: {str(e)}")
        
        return '\n'.join(text) if text else "Não foi possível extrair texto via OCR"
    except Exception as e:
        logger.error(f"Erro no OCR: {str(e)}\n{traceback.format_exc()}")
        return "Erro crítico no processamento OCR"

def extract_text(file):
    """Extrai texto de diferentes formatos de arquivo"""
    try:
        content = file.read()
        filename = file.filename
        
        if filename.endswith('.pdf'):
            try:
                pdf = PdfReader(io.BytesIO(content))
                text = '\n'.join([page.extract_text() for page in pdf.pages])
                if text.strip():
                    return text
            except Exception as e:
                logger.error(f"Erro na extração PDF: {str(e)}")
            
            return extract_text_with_ocr(content)
        
        elif filename.endswith('.docx'):
            doc = Document(io.BytesIO(content))
            return '\n'.join([para.text for para in doc.paragraphs])
        
        elif filename.endswith(('.xlsx', '.xls')):
            wb = load_workbook(io.BytesIO(content))
            text = []
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    text.append(' '.join(map(str, row)))
            return '\n'.join(text)
            
    except Exception as e:
        logger.error(f"Erro na extração: {str(e)}\n{traceback.format_exc()}")
        return "Erro na extração do texto"

def generate_summary(text, service, api_key):
    """Gera o resumo usando o serviço de IA selecionado"""
    try:
        prompt = f"Resuma este documento em português brasileiro de forma clara e detalhada:\n\n{text}"
        
        if service == 'OpenAI':
            headers = {'Authorization': f'Bearer {api_key}'}
            data = {
                "model": "gpt-3.5-turbo",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.7
            }
            response = requests.post(
                AI_SERVICES[service]['api_url'],
                json=data,
                headers=headers,
                timeout=60
            )
            return response.json()['choices'][0]['message']['content']
        
        elif service == 'HuggingFace':
            headers = {'Authorization': f'Bearer {api_key}'}
            data = {"inputs": prompt}
            response = requests.post(
                AI_SERVICES[service]['api_url'],
                json=data,
                headers=headers,
                timeout=120
            )
            return response.json()[0]['generated_text']
        
        elif service == 'Cohere':
            headers = {
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            }
            data = {
                "prompt": prompt,
                "model": "command",
                "max_tokens": 1000,
                "temperature": 0.7
            }
            response = requests.post(
                AI_SERVICES[service]['api_url'],
                json=data,
                headers=headers,
                timeout=60
            )
            return response.json()['generations'][0]['text']
    
    except Exception as e:
        logger.error(f"Erro na geração do resumo: {str(e)}\n{traceback.format_exc()}")
        return f"Erro ao gerar resumo: {str(e)}"

@app.errorhandler(500)
def handle_500(error):
    logger.error(f"Erro 500: {error}\n{traceback.format_exc()}")
    return render_template_string(HTML_BASE, content='''
        <div class="alert alert-danger">
            <h4>Erro Interno</h4>
            <p>Por favor, tente novamente ou verifique os logs</p>
            <a href="/" class="btn btn-secondary">Voltar</a>
        </div>
    '''), 500

@app.route('/')
def home():
    return render_template_string(HTML_BASE, content='''
        <div class="text-center">
            <h4 class="mb-4">Selecione uma opção:</h4>
            <a href="/settings" class="btn btn-primary btn-lg mb-2">⚙️ Configurações</a><br>
            <a href="/process" class="btn btn-success btn-lg">📄 Processar Documento</a>
        </div>
    ''')

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'POST':
        session['ai_service'] = request.form.get('ai_service')
        return redirect(f'/configure/{request.form.get("ai_service")}')
    
    return render_template_string(HTML_BASE, content='''
        <h4 class="mb-4">⚙️ Configurações</h4>
        <form method="POST">
            <div class="mb-3">
                <label>Serviço de IA:</label>
                <select name="ai_service" class="form-select" required>
                    <option value="">Selecione...</option>
                    <option value="OpenAI">OpenAI</option>
                    <option value="HuggingFace">HuggingFace</option>
                    <option value="Cohere">Cohere</option>
                </select>
            </div>
            <button type="submit" class="btn btn-primary">Continuar</button>
        </form>
    ''')

@app.route('/configure/<service>', methods=['GET', 'POST'])
def configure(service):
    if request.method == 'POST':
        session['api_key'] = request.form.get('api_key')
        return redirect('/')
    
    return render_template_string(HTML_BASE, content=f'''
        <h4 class="mb-4">🔧 Configurar {service}</h4>
        <div class="mb-4">
            {'<hr>'.join(f'<div class="guide-step">{step}</div>' for step in AI_SERVICES[service]['guide'])}
        </div>
        <form method="POST">
            <div class="mb-3">
                <label>Chave API:</label>
                <input type="text" name="api_key" class="form-control" required>
            </div>
            <button type="submit" class="btn btn-primary">Salvar</button>
        </form>
    ''')

@app.route('/process', methods=['GET', 'POST'])
def process():
    try:
        if 'api_key' not in session:
            return redirect('/settings')
        
        if request.method == 'POST':
            if 'file' in request.files:
                file = request.files['file']
                if file.filename == '':
                    return redirect(request.url)
                
                extracted_text = extract_text(file)
                session['extracted_text'] = extracted_text
                session['filename'] = file.filename
                
                return render_template_string(HTML_BASE, content=f'''
                    <h4 class="mb-4">✏️ Editar Texto</h4>
                    <form method="POST">
                        <div class="mb-3">
                            <label>Texto extraído de {file.filename}:</label>
                            <textarea name="edited_text" class="form-control">{extracted_text}</textarea>
                        </div>
                        <button type="submit" class="btn btn-primary">Gerar Resumo</button>
                    </form>
                ''')
            else:
                edited_text = request.form.get('edited_text', '')
                summary = generate_summary(
                    edited_text,
                    session['ai_service'],
                    session['api_key']
                )
                
                return render_template_string(HTML_BASE, content=f'''
                    <div class="alert alert-info">
                        <h5>Texto Enviado:</h5>
                        <pre>{edited_text[:2000]}{'...' if len(edited_text) > 2000 else ''}</pre>
                    </div>
                    <div class="alert alert-success">
                        <h5>Resumo Gerado:</h5>
                        <pre>{summary}</pre>
                    </div>
                    <a href="/process" class="btn btn-primary">Nova Análise</a>
                ''')
        
        return render_template_string(HTML_BASE, content='''
            <h4 class="mb-4">📤 Enviar Documento</h4>
            <form method="POST" enctype="multipart/form-data">
                <div class="mb-3">
                    <label>Selecione o arquivo (PDF/DOCX/XLSX):</label>
                    <input type="file" name="file" class="form-control" required>
                </div>
                <button type="submit" class="btn btn-primary">Enviar</button>
            </form>
        ''')
    
    except Exception as e:
        logger.error(f"Erro no processamento: {str(e)}\n{traceback.format_exc()}")
        return render_template_string(HTML_BASE, content=f'''
            <div class="alert alert-danger">
                <h4>Erro no Processamento</h4>
                <pre>{str(e)}</pre>
                <a href="/" class="btn btn-secondary">Voltar</a>
            </div>
        '''), 500

if __name__ == '__main__':
    nltk.download('stopwords', quiet=True)
    nltk.download('punkt', quiet=True)
    nltk.download('wordnet', quiet=True)
    nltk.download('omw-1.4', quiet=True)
    app.run(debug=False)

if __name__ == '__main__':
    app.run(debug=True)
